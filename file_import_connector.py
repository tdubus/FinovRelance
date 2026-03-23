"""
File Import Connector for Excel/CSV imports
Supports .xlsx and .csv file formats with custom column mapping

This module provides utilities for:
- Detecting column headers from Excel/CSV files
- Transforming files according to saved mappings
- Converting to standard format for import_views.py
"""

import io
import logging
from typing import List, Dict, Tuple, Optional
from datetime import datetime, date
from openpyxl import load_workbook
from defusedcsv import csv

logger = logging.getLogger(__name__)


ADDRESS_FIELDS = ['address_line1', 'address_line2', 'city', 'province', 'postal_code', 'country']


def _concatenate_address_parts(parts: Dict[str, str]) -> str:
    """
    Concatenate address parts into a single address string.

    Args:
        parts: Dictionary with keys: address_line1, address_line2, city, province, postal_code, country

    Returns:
        Concatenated address string like: "247 rue test, Saint-Test, QC, J0L 1Y0, Canada"
    """
    address_components = []

    # Add address lines (combine line1 and line2 if both exist)
    line1 = parts.get('address_line1', '').strip()
    line2 = parts.get('address_line2', '').strip()
    if line1 and line2:
        address_components.append(f"{line1}, {line2}")
    elif line1:
        address_components.append(line1)
    elif line2:
        address_components.append(line2)

    # Add city
    city = parts.get('city', '').strip()
    if city:
        address_components.append(city)

    # Add province
    province = parts.get('province', '').strip()
    if province:
        address_components.append(province)

    # Add postal code
    postal_code = parts.get('postal_code', '').strip()
    if postal_code:
        address_components.append(postal_code)

    # Add country
    country = parts.get('country', '').strip()
    if country:
        address_components.append(country)

    return ', '.join(address_components)


def _map_language_value(value: str, language_mappings: Optional[Dict[str, str]]) -> str:
    """
    Map a language value from the file to internal code (FR or EN)

    Args:
        value: Language value from the file (e.g., "Français", "French", "FR", "EN")
        language_mappings: Dictionary mapping codes to values (e.g., {"FR": "Français", "EN": "Anglais"})

    Returns:
        Internal language code (normalized to lowercase)
        - If a mapping exists and matches, returns the mapped code (e.g., "Français" → "fr")
        - Otherwise, returns the original value normalized (e.g., "FR" → "fr", "EN" → "en")
        - Defaults to 'fr' if value is empty
    """
    if not value:
        return 'fr'

    value_normalized = value.strip().lower()

    # If mappings are provided, try to find a match
    if language_mappings:
        for code, mapped_value in language_mappings.items():
            if mapped_value.strip().lower() == value_normalized:
                return code.lower()

    # If no mapping matched, return the original value normalized
    # This allows direct codes like "FR", "EN", "fr", "en" to work
    return value_normalized


def detect_file_type(filename: str) -> str:
    """
    Detect file type from filename extension

    Args:
        filename: Name of the file

    Returns:
        'excel' or 'csv' or 'unknown'
    """
    filename_lower = filename.lower()
    if filename_lower.endswith('.xlsx') or filename_lower.endswith('.xls'):
        return 'excel'
    elif filename_lower.endswith('.csv'):
        return 'csv'
    else:
        return 'unknown'


def detect_headers_from_file(file_content: bytes, file_type: str) -> Tuple[List[str], Optional[str]]:
    """
    Detect column headers from Excel or CSV file

    Args:
        file_content: Raw file content as bytes
        file_type: 'excel' or 'csv'

    Returns:
        Tuple of (headers_list, error_message)
        headers_list is empty if error occurred
    """
    try:
        if file_type == 'excel':
            return _detect_headers_excel(file_content)
        elif file_type == 'csv':
            return _detect_headers_csv(file_content)
        else:
            return [], "Type de fichier non supporté. Utilisez .xlsx ou .csv"
    except Exception as e:
        logger.error(f"Error detecting headers: {e}")
        return [], f"Erreur lors de la lecture du fichier : {str(e)}"


def _detect_headers_excel(file_content: bytes) -> Tuple[List[str], Optional[str]]:
    """Detect headers from Excel file"""
    try:
        # Load workbook from bytes
        workbook = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
        sheet = workbook.active

        # Read first row as headers
        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).strip())
            else:
                headers.append(f"Colonne_{len(headers)+1}")

        workbook.close()

        if not headers:
            return [], "Aucune colonne détectée dans le fichier Excel"

        logger.info(f"Detected {len(headers)} headers from Excel: {headers}")
        return headers, None

    except Exception as e:
        logger.error(f"Error reading Excel file: {e}")
        return [], f"Erreur lecture Excel : {str(e)}"


def _detect_headers_csv(file_content: bytes) -> Tuple[List[str], Optional[str]]:
    """Detect headers from CSV file"""
    try:
        # Decode bytes to string
        content = file_content.decode('utf-8')

        # Normalize line endings
        content = content.replace('\r\n', '\n').replace('\r', '\n')

        # Parse CSV
        stream = io.StringIO(content)
        csv_reader = csv.reader(stream, delimiter=',', quotechar='"')

        # Read first row as headers
        headers = next(csv_reader, None)

        if not headers:
            return [], "Aucune colonne détectée dans le fichier CSV"

        # Clean headers
        headers = [str(h).strip() for h in headers]

        logger.info(f"Detected {len(headers)} headers from CSV: {headers}")
        return headers, None

    except Exception as e:
        logger.error(f"Error reading CSV file: {e}")
        return [], f"Erreur lecture CSV : {str(e)}"


def transform_file_to_standard_format(
    file_content: bytes,
    file_type: str,
    mapping: Dict[str, str],
    import_type: str,
    language_mappings: Optional[Dict[str, str]] = None,
    include_project_field: bool = False
) -> Tuple[List[List[str]], int, List[str]]:
    """
    Transform Excel/CSV file according to mapping configuration

    Args:
        file_content: Raw file content as bytes
        file_type: 'excel' or 'csv'
        mapping: Dictionary mapping file columns to standard fields
                 Format: {"column_in_file": "standard_field_name"}
        import_type: 'clients' or 'invoices'
        language_mappings: Dictionary mapping language values to codes
                          Format: {"FR": "Français", "EN": "Anglais"}
        include_project_field: If True, includes project_name field for invoices

    Returns:
        Tuple of (rows_data, total_rows, errors)
        rows_data: List of lists representing transformed rows
        total_rows: Number of data rows processed
        errors: List of error messages
    """
    try:
        if file_type == 'excel':
            return _transform_excel_file(file_content, mapping, import_type, language_mappings, include_project_field)
        elif file_type == 'csv':
            return _transform_csv_file(file_content, mapping, import_type, language_mappings, include_project_field)
        else:
            return [], 0, ["Type de fichier non supporté"]
    except Exception as e:
        logger.error(f"Error transforming file: {e}")
        return [], 0, [f"Erreur transformation : {str(e)}"]


def _convert_excel_cell_value(value) -> str:
    """
    Convert Excel cell value to string, handling dates properly

    Args:
        value: Cell value from openpyxl (can be datetime, str, int, float, etc.)

    Returns:
        String representation of the value
    """
    if value is None:
        return ''

    # Handle datetime objects - convert to YYYY-MM-DD format
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')

    # Handle date objects - convert to YYYY-MM-DD format
    if isinstance(value, date):
        return value.strftime('%Y-%m-%d')

    # For all other types (str, int, float, etc.), convert to string and strip
    return str(value).strip()


def _transform_excel_file(
    file_content: bytes,
    mapping: Dict[str, str],
    import_type: str,
    language_mappings: Optional[Dict[str, str]] = None,
    include_project_field: bool = False
) -> Tuple[List[List[str]], int, List[str]]:
    """Transform Excel file according to mapping"""
    try:
        workbook = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
        sheet = workbook.active

        # Read headers from first row
        headers = []
        for cell in sheet[1]:
            headers.append(str(cell.value).strip() if cell.value else "")

        # Get standard field order based on import type
        standard_fields = _get_standard_fields(import_type, include_project_field)

        # Create column mapping (file column index -> standard field index)
        # Also track address field mappings separately
        column_mapping = {}
        address_column_mapping = {}  # file_col_index -> address_field_name

        for file_col, standard_field in mapping.items():
            try:
                file_col_index = headers.index(file_col)

                # Check if this is an address field
                if standard_field in ADDRESS_FIELDS:
                    address_column_mapping[file_col_index] = standard_field
                else:
                    standard_field_index = standard_fields.index(standard_field)
                    column_mapping[file_col_index] = standard_field_index
            except ValueError:
                logger.warning(f"Column '{file_col}' not found in file or standard fields")

        # Find language and address field indices if they exist
        language_field_index = None
        if 'language' in standard_fields:
            language_field_index = standard_fields.index('language')

        address_field_index = None
        if 'address' in standard_fields:
            address_field_index = standard_fields.index('address')

        # Transform data rows
        transformed_rows = []
        errors = []

        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Create empty standard row
                standard_row = [''] * len(standard_fields)

                # Collect address parts for concatenation
                address_parts = {}

                # Map values from file to standard positions
                for file_idx, standard_idx in column_mapping.items():
                    if file_idx < len(row) and row[file_idx] is not None:
                        value = _convert_excel_cell_value(row[file_idx])

                        # Apply language mapping if this is the language field
                        if standard_idx == language_field_index and language_mappings:
                            value = _map_language_value(value, language_mappings)

                        standard_row[standard_idx] = value

                # Collect address field values
                for file_idx, addr_field in address_column_mapping.items():
                    if file_idx < len(row) and row[file_idx] is not None:
                        address_parts[addr_field] = _convert_excel_cell_value(row[file_idx])

                # Concatenate address fields if any were provided
                if address_parts and address_field_index is not None:
                    concatenated_address = _concatenate_address_parts(address_parts)
                    standard_row[address_field_index] = concatenated_address

                # Only add row if it has at least one non-empty value
                if any(standard_row):
                    transformed_rows.append(standard_row)

            except Exception as e:
                errors.append(f"Ligne {row_num}: {str(e)}")

        workbook.close()

        logger.info(f"Transformed {len(transformed_rows)} rows from Excel")
        return transformed_rows, len(transformed_rows), errors

    except Exception as e:
        logger.error(f"Error transforming Excel: {e}")
        return [], 0, [f"Erreur transformation Excel : {str(e)}"]


def _transform_csv_file(
    file_content: bytes,
    mapping: Dict[str, str],
    import_type: str,
    language_mappings: Optional[Dict[str, str]] = None,
    include_project_field: bool = False
) -> Tuple[List[List[str]], int, List[str]]:
    """Transform CSV file according to mapping"""
    try:
        # Decode and normalize
        content = file_content.decode('utf-8')
        content = content.replace('\r\n', '\n').replace('\r', '\n')

        stream = io.StringIO(content)
        csv_reader = csv.reader(stream, delimiter=',', quotechar='"')

        # Read headers
        headers = next(csv_reader, None)
        if not headers:
            return [], 0, ["Fichier CSV vide"]

        headers = [h.strip() for h in headers]

        # Get standard field order
        standard_fields = _get_standard_fields(import_type, include_project_field)

        # Create column mapping
        # Also track address field mappings separately
        column_mapping = {}
        address_column_mapping = {}  # file_col_index -> address_field_name

        for file_col, standard_field in mapping.items():
            try:
                file_col_index = headers.index(file_col)

                # Check if this is an address field
                if standard_field in ADDRESS_FIELDS:
                    address_column_mapping[file_col_index] = standard_field
                else:
                    standard_field_index = standard_fields.index(standard_field)
                    column_mapping[file_col_index] = standard_field_index
            except ValueError:
                logger.warning(f"Column '{file_col}' not found in CSV or standard fields")

        # Find language and address column indices if they exist
        language_field_index = None
        if 'language' in standard_fields:
            language_field_index = standard_fields.index('language')

        address_field_index = None
        if 'address' in standard_fields:
            address_field_index = standard_fields.index('address')

        # Transform data rows
        transformed_rows = []
        errors = []

        for row_num, row in enumerate(csv_reader, start=2):
            try:
                # Create empty standard row
                standard_row = [''] * len(standard_fields)

                # Collect address parts for concatenation
                address_parts = {}

                # Map values
                for file_idx, standard_idx in column_mapping.items():
                    if file_idx < len(row):
                        value = row[file_idx].strip()

                        # Apply language mapping if this is the language field
                        if standard_idx == language_field_index and language_mappings:
                            value = _map_language_value(value, language_mappings)

                        standard_row[standard_idx] = value

                # Collect address field values
                for file_idx, addr_field in address_column_mapping.items():
                    if file_idx < len(row):
                        address_parts[addr_field] = row[file_idx].strip()

                # Concatenate address fields if any were provided
                if address_parts and address_field_index is not None:
                    concatenated_address = _concatenate_address_parts(address_parts)
                    standard_row[address_field_index] = concatenated_address

                # Only add non-empty rows
                if any(standard_row):
                    transformed_rows.append(standard_row)

            except Exception as e:
                errors.append(f"Ligne {row_num}: {str(e)}")

        logger.info(f"Transformed {len(transformed_rows)} rows from CSV")
        return transformed_rows, len(transformed_rows), errors

    except Exception as e:
        logger.error(f"Error transforming CSV: {e}")
        return [], 0, [f"Erreur transformation CSV : {str(e)}"]


def _get_standard_fields(import_type: str, include_project_field: bool = False) -> List[str]:
    """
    Get standard field order for import type
    Must match the column order expected by import_views.py

    Args:
        import_type: 'clients' or 'invoices'
        include_project_field: If True, adds project_name field for invoices (feature flag)

    Returns:
        List of standard field names in expected order
    """
    if import_type == 'clients':
        # Order: code_client, name, email, phone, address, representative_name, payment_terms, parent_code, language
        return [
            'code_client',
            'name',
            'email',
            'phone',
            'address',
            'representative_name',
            'payment_terms',
            'parent_code',
            'language'
        ]
    elif import_type == 'invoices':
        # Order: code_client, invoice_number, amount, original_amount, issue_date, due_date[, project_name]
        fields = [
            'code_client',
            'invoice_number',
            'amount',
            'original_amount',
            'issue_date',
            'due_date'
        ]
        if include_project_field:
            fields.append('project_name')
        return fields
    else:
        return []


def get_required_fields(import_type: str) -> List[str]:
    """
    Get list of required fields for import type

    Args:
        import_type: 'clients' or 'invoices'

    Returns:
        List of required field names
    """
    if import_type == 'clients':
        return ['code_client', 'name']
    elif import_type == 'invoices':
        return ['code_client', 'invoice_number', 'amount', 'issue_date', 'due_date']
    else:
        return []


def validate_mapping(mapping: Dict[str, str], import_type: str) -> Tuple[bool, List[str]]:
    """
    Validate that mapping includes all required fields

    Args:
        mapping: Dictionary mapping file columns to standard fields
        import_type: 'clients' or 'invoices'

    Returns:
        Tuple of (is_valid, error_messages)
    """
    required_fields = get_required_fields(import_type)
    mapped_fields = set(mapping.values())

    errors = []
    for required in required_fields:
        if required not in mapped_fields:
            errors.append(f"Champ obligatoire manquant : {required}")

    is_valid = len(errors) == 0
    return is_valid, errors
